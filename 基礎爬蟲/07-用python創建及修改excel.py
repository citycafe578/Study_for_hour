from openpyxl import Workbook
import openpyxl

wb = Workbook()
ws = wb.active
ws.title = "創建excel檔案練習"

#依次輸入一整行內容
title = ["姓名","座號","成績"]
ws.append(title)
ws.append(["CT",1,100])
ws.append(["老劉",2,85])
ws.append(["彥字",3,92])
ws.append(["使用逗號","就可以","在原有的一行中","向右新增資料"])
ws["A6"] = "也可以指定位置" 

#顯示最大排數
print("直排總數:", ws.max_row)
print("橫排總數:", ws.max_column)

#逐一顯示內容
for i in range(1, ws.max_row + 1, 1):
    for j in range(1, ws.max_column + 1, 1):
        print(ws.cell(row = i, column=j).value)

wb.save("新excel檔案.xlsx")